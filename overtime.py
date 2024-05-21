import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

# Excel文書を作成する関数
def create_excel(data):
    wb = openpyxl.load_workbook('template.xlsx')
    ws = wb.active

    # データをExcelシートに追加
    ws["D10"] = data["申請日"]
    ws["D10"].alignment = Alignment(horizontal='center', vertical='center')
    ws["D24"].alignment = Alignment(horizontal='center', vertical='center')
    ws["E12"] = data["職種"]
    ws["E12"].alignment = Alignment(horizontal='center', vertical='center')
    ws["E26"].alignment = Alignment(horizontal='center', vertical='center')
    ws["D10"].alignment = Alignment(horizontal='center', vertical='center')
    ws["E13"] = data["氏名"]
    ws["E13"].alignment = Alignment(horizontal='center', vertical='center')
    ws["E27"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B4"] = data["期日"]
    ws["B4"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B18"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B5"] = data["開始時間"]
    ws["B5"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B19"].alignment = Alignment(horizontal='center', vertical='center')
    ws["E5"] = data["終了時間"]
    ws["E5"].alignment = Alignment(horizontal='center', vertical='center')
    ws["F19"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B6"] = data["事由"]
    ws["B6"].alignment = Alignment(horizontal='center', vertical='center')
    ws["B20"].alignment = Alignment(horizontal='center', vertical='center')

    # 開始時間と終了時間の差を計算して分形式でG5に出力
    start_time = data["開始時間_obj"]
    end_time = data["終了時間_obj"]
    duration = end_time - start_time
    duration_minutes = duration.total_seconds() / 60
    ws["G5"] = f"{int(duration_minutes)}"
    ws["G5"].alignment = Alignment(horizontal='center', vertical='center')
    
    # ファイル名を生成
    filename = f"残業届_{data['申請日']}_{data['職種']}_{data['氏名']}.xlsx"
    wb.save(filename)
    return filename

# Streamlit UI
st.title("業務改善アプリ")
st.header("残業申請")

# ユーザー入力を収集
name = st.text_input("氏名")
request_date = st.date_input("申請日")
formatted_request_date = f"令和{request_date.year - 2018}年{request_date.month}月{request_date.day}日"
department = st.selectbox("職種", ["医師", "看護師", "介護士", "ケアマネジャー", "給食", "事務", "受付", "営繕"])
start_date = st.date_input("期日")
formatted_start_date = f"令和{start_date.year - 2018}年{start_date.month}月{start_date.day}日"
start_time = st.time_input("開始時間")
formatted_start_time = f"{start_time.hour}時{start_time.minute}分"
end_time = st.time_input("終了時間")
formatted_end_time = f"{end_time.hour}時{end_time.minute}分"
reason = st.text_area("事由")

# 申請ボタン
if st.button("申請書作成"):
    data = {
        "申請日": formatted_request_date,
        "職種": department,
        "期日": formatted_start_date,
        "氏名": name,
        "開始時間": formatted_start_time,
        "終了時間": formatted_end_time,
        "開始時間_obj": datetime.combine(datetime.today(), start_time),
        "終了時間_obj": datetime.combine(datetime.today(), end_time),
        "事由": reason,
    }
    filename = create_excel(data)
    st.success("申請書が作成されました。")

    # ダウンロードリンクの提供
    with open(filename, "rb") as file:
        st.download_button("申請書をダウンロード", data=file, file_name=filename)


